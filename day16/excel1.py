import openpyxl
from random import uniform

planilha = openpyxl.Workbook()
planilha.create_sheet('Sheet1', 0)
planilha.create_sheet('Sheet2', 1)

planilha1 = planilha['Sheet1']
planilha2 = planilha['Sheet2']

for line in range(1, 10):
    request_number = line-1
    planilha1.cell(line, 1).value = request_number
    planilha1.cell(line, 2).value = 1200 + (line - 4)
    price = round(uniform(10, 1000), 2)
    planilha1.cell(line, 3).value = price

for line in range(1, 15):
    planilha2.cell(line, 1).value = f'Vinicius{line} {round(uniform(10, 100), 2)}'
    planilha2.cell(line, 2).value = f'Maria{line} {round(uniform(10, 100), 2)}'
    planilha2.cell(line, 3).value = f'Betânia{line} {round(uniform(10, 100), 2)}'

planilha.save('nova_planilha.xlsx')


# requestions = openpyxl.load_workbook('/home/vinicius/Documents/udemyPython/day16/pedidos.xlsx')
# sheet_names = requestions.sheetnames
# sheet1 = requestions['Página1']
# for line in range(5, 16):
#     request_number = line-1
#     sheet1.cell(line, 1).value = request_number
#     sheet1.cell(line, 2).value = 1200 + (line - 4)
#     price = round(uniform(10, 1000), 2)
#     sheet1.cell(line, 3).value = price
#     print(line)
#
#
# requestions.save('new_plane.xlsx')







# for line in sheet1:
#     print(line[0].value, line[1].value, line[2].value, line[3].value)
#





# for line in sheet1['a1:c2']:
#     for columns in line:
#         print(columns.value)




# for field in sheet1['b']:
#     print(field.value)
# print(sheet1['b4'].value)

